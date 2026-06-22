import base64
import io
import json
import uuid
from datetime import date
from typing import Optional

from groq import Groq
from sqlalchemy.orm import Session

from app.core.config import settings
from app.models.student_progress import StudentProgress, ProgressStatus
from app.models.student_requirement import StudentRequirement
from app.models.student_stage_deadline import StudentStageDeadline
from app.models.stage import Stage
from app.models.requirement import Requirement
from app.models.university import University
from app.models.user import User, AccountStatus

# Основная модель — точная 70b; при исчерпании её дневного лимита (429) авто-фолбэк на 8b.
TEXT_MODEL = "llama-3.3-70b-versatile"
FALLBACK_MODEL = "llama-3.1-8b-instant"
VISION_MODEL = "meta-llama/llama-4-scout-17b-16e-instruct"  # актуальная мультимодальная модель Groq


def _gpa_percent(gpa: Optional[float]) -> int:
    """Normalize a GPA to a 0-100 score, auto-detecting the source scale.

    Supports the common 4.0, 5.0, 10 and 100-point scales."""
    if not gpa or gpa <= 0:
        return 0
    if gpa <= 4.0:
        scale = 4.0
    elif gpa <= 5.0:
        scale = 5.0
    elif gpa <= 10.0:
        scale = 10.0
    else:
        scale = 100.0
    return min(100, int(gpa / scale * 100))


import logging
import time

logger = logging.getLogger("ai_service")

GROQ_TIMEOUT = 30  # seconds — avoid a slow upstream tying up a worker thread
GROQ_RETRIES = 3   # повтор при временных сбоях Groq (rate limit / таймаут / 5xx)


def _client() -> Groq:
    if not settings.GROQ_API_KEY:
        raise ValueError("GROQ_API_KEY не задан в настройках")
    return Groq(api_key=settings.GROQ_API_KEY, timeout=GROQ_TIMEOUT)


def _complete(messages: list, model: str = TEXT_MODEL, max_tokens: int = 2048) -> str:
    """Вызов Groq с ретраями и авто-фолбэком модели при исчерпании лимита (429)."""
    # Цепочка моделей: основная → запасная (для текста). Vision — без фолбэка.
    if model == VISION_MODEL:
        chain = [model]
    else:
        chain = [model] + ([FALLBACK_MODEL] if FALLBACK_MODEL != model else [])

    last_err: Exception | None = None
    for m in chain:
        for attempt in range(GROQ_RETRIES):
            try:
                resp = _client().chat.completions.create(
                    model=m,
                    messages=messages,
                    max_tokens=max_tokens,
                    temperature=0.7,
                )
                return resp.choices[0].message.content.strip()
            except Exception as e:  # noqa: BLE001
                last_err = e
                status = getattr(e, "status_code", None)
                logger.warning("Groq %s failed (attempt %s/%s, status=%s): %s",
                               m, attempt + 1, GROQ_RETRIES, status, e)
                if status == 429:
                    break  # дневной/минутный лимит модели — сразу к запасной модели
                if attempt < GROQ_RETRIES - 1 and (status in (408, 500, 502, 503, 504) or status is None):
                    time.sleep(0.8 * (attempt + 1))
                    continue
                break
    raise last_err if last_err else RuntimeError("Groq call failed")


def _extract_json(text: str, opener: str, closer: str) -> str:
    """Pull the first balanced JSON object/array out of an LLM reply.

    Handles ```json fences and any leading/trailing prose the model adds."""
    text = text.strip()
    if "```" in text:
        # Take the content of the first fenced block
        parts = text.split("```")
        if len(parts) >= 2:
            block = parts[1]
            if block.lower().startswith("json"):
                block = block[4:]
            text = block.strip()
    start = text.find(opener)
    end = text.rfind(closer)
    if start != -1 and end != -1 and end > start:
        return text[start:end + 1]
    return text


def _parse_json(text: str, fallback: dict) -> dict:
    try:
        result = json.loads(_extract_json(text, "{", "}"))
        return result if isinstance(result, dict) else fallback
    except Exception:
        return fallback


def _parse_json_list(text: str) -> list:
    try:
        result = json.loads(_extract_json(text, "[", "]"))
        return result if isinstance(result, list) else []
    except Exception:
        return []


# ── Context builder ───────────────────────────────────────────────────────────

def _build_student_context(db: Session, user: User) -> str:
    lines = [
        f"Студент: {user.full_name}",
        f"GPA: {user.gpa or 'не указан'}",
        f"IELTS: {user.ielts_score or 'не указан'}",
        f"TOEFL: {user.toefl_score or 'не указан'}",
        f"SAT: {user.sat_score or 'не указан'}",
        f"HSK: {f'уровень {user.hsk_level}' if user.hsk_level else 'не сдавал'}",
        f"Гражданство: {user.citizenship or 'не указано'}",
        f"Желаемая специальность: {user.desired_specialty or 'не указана'}",
        f"Предпочтительные страны: {', '.join(user.country_preference) if user.country_preference else 'не указаны'}",
        f"Уровень программы: {user.program_level or 'не указан'}",
        f"Языковой год: {user.wants_language_year or 'не указан'}",
        f"Предпочитаемая сложность: {user.preferred_difficulty or 'не указана'}",
        f"Макс. бюджет (RMB/год): {user.max_budget_rmb or 'не указан'}",
        f"Достижения: {user.achievements or 'не указаны'}",
    ]

    progress = db.query(StudentProgress).filter(
        StudentProgress.user_id == user.id,
        StudentProgress.status == ProgressStatus.in_progress,
    ).first()

    if progress:
        uni = db.get(University, progress.university_id)
        if uni:
            lines.append(f"Выбранный университет: {uni.name} ({uni.country}, {uni.city}), рейтинг: {uni.rating or 'нет'}, стоимость: ${uni.cost or 'нет'}/год")

        if progress.current_stage_id:
            stage = db.get(Stage, progress.current_stage_id)
            if stage:
                lines.append(f"Текущий этап: {stage.name} (этап №{stage.order})")

                ssd = db.query(StudentStageDeadline).filter(
                    StudentStageDeadline.student_progress_id == progress.id,
                    StudentStageDeadline.stage_id == stage.id,
                ).first()
                if ssd:
                    days_left = (ssd.deadline - date.today()).days
                    if days_left > 0:
                        lines.append(f"Дней до дедлайна: {days_left}")
                    else:
                        lines.append(f"⚠️ Дедлайн просрочен на {abs(days_left)} дней!")

                reqs = db.query(Requirement).filter(Requirement.stage_id == stage.id).all()
                # One query for all completed requirements of this progress (avoids N+1)
                done_req_ids = {
                    row[0] for row in db.query(StudentRequirement.requirement_id).filter(
                        StudentRequirement.student_progress_id == progress.id,
                        StudentRequirement.completed == True,  # noqa: E712
                    ).all()
                }
                unfulfilled = [req.name for req in reqs if req.id not in done_req_ids]
                if unfulfilled:
                    lines.append(f"Не выполненные требования: {', '.join(unfulfilled)}")
                else:
                    lines.append("Все требования текущего этапа выполнены.")
    else:
        lines.append("Студент ещё не выбрал университет для поступления.")

    return "\n".join(lines)


# ── Барашек: постоянный гид-ассистент по системе ──────────────────────────────

# Описание каждой страницы кабинета: что это, какие кнопки и что делать.
# Барашек использует это, чтобы подсказывать «куда нажимать».
BARASHEK_PAGES = {
    "training": (
        "Страница «Мой путь» (/dashboard/training) — здесь видны этапы поступления по "
        "выбранному университету. На каждом этапе есть чеклист требований (галочки), "
        "загрузка документов и видео-уроки. Кнопки: отметить требование выполненным, "
        "загрузить файл, посмотреть урок, перейти к следующему этапу."
    ),
    "universities": (
        "Страница «Университеты» (/dashboard/universities) — каталог вузов. "
        "Вверху справа кнопки: «Подобрать по моим данным» (AI-подбор под профиль) и "
        "«Сравнить». На карточке вуза: кнопка «Подробнее», иконка-сердечко добавляет в "
        "Избранное. Есть поиск и фильтр по стране."
    ),
    "favourites": (
        "Страница «Избранное» (/dashboard/favourites) — сохранённые вузы. Отсюда можно "
        "открыть вуз и начать по нему путь поступления."
    ),
    "notifications": (
        "Страница «Уведомления» (/dashboard/notifications) — напоминания о дедлайнах и "
        "новых этапах. Клик по уведомлению отмечает его прочитанным."
    ),
    "profile": (
        "Страница «Профиль» (/dashboard/profile) — личные данные, GPA, баллы тестов "
        "(IELTS/TOEFL/SAT/HSK), желаемая специальность и бюджет. Здесь всё можно "
        "отредактировать и сохранить. Полный профиль нужен для точного AI-подбора."
    ),
    "university_detail": (
        "Страница конкретного университета — подробная информация, стоимость, "
        "специальности, требования. Здесь можно добавить вуз в избранное и начать "
        "по нему путь поступления."
    ),
}


def barashek_guide(db: Session, user: User, message: str, history: list,
                   page: str = "", auto: bool = False) -> str:
    """Постоянный гид-ассистент «Барашек».

    Тепло ведёт студента по системе: подсказывает куда нажимать и что делать,
    отвечает форматированным markdown, хвалит через раз. Если auto=True —
    проактивно объясняет текущую страницу без вопроса студента.
    """
    context = _build_student_context(db, user)
    page_info = BARASHEK_PAGES.get(page, "")

    # «Хвалит через 1 раз»: считаем уже сказанные ответы Барашка.
    assistant_count = sum(1 for h in history if h.get("role") == "assistant")
    should_praise = (assistant_count % 2 == 0)

    system = (
        "Ты — Барашек 🐑, тёплый, заботливый AI-проводник студента в личном кабинете "
        "платформы EduBridge University (поступление в зарубежные вузы).\n\n"
        "ТВОЯ РОЛЬ:\n"
        "- Ты постоянный помощник: ведёшь студента за руку по системе, "
        "конкретно подсказываешь КУДА НАЖИМАТЬ и ЧТО ДЕЛАТЬ прямо сейчас.\n"
        "- Ссылайся на реальные кнопки и разделы по их названиям (например: "
        "«нажми кнопку **Подобрать по моим данным** вверху справа»).\n"
        "- Тон тёплый, ласковый, с эмодзи 🐑✨💚, обращайся по имени.\n"
        "- Отвечай ТОЛЬКО на русском.\n"
        "- ВСЕГДА форматируй ответ в markdown: **жирным** выделяй кнопки и важное, "
        "используй короткие списки шагов (1. 2. 3. или •). Будь краток — 2-5 пунктов.\n"
        + (
            "- ВАЖНО: в этом ответе ОБЯЗАТЕЛЬНО искренне похвали и подбодри студента "
            "(«умничка», «ты молодец», «горжусь тобой») в начале.\n"
            if should_praise else
            "- В этом ответе НЕ хвали — сразу по делу, по-доброму.\n"
        )
        + f"\nИмя студента: {user.full_name}\n"
        f"\nПрофиль и прогресс студента:\n{context}\n"
        + (f"\nСтудент сейчас на странице:\n{page_info}\n" if page_info else "")
    )

    if auto:
        user_msg = (
            "Студент только что открыл эту страницу. Кратко и тепло поприветствуй его "
            "здесь, объясни что это за раздел и подскажи 1-3 конкретных шага, что делать "
            "дальше. Укажи на нужные кнопки."
        )
    else:
        user_msg = message

    messages = [{"role": "system", "content": system}]
    for h in history[-10:]:
        if h.get("role") in ("user", "assistant"):
            messages.append({"role": h["role"], "content": h["content"]})
    messages.append({"role": "user", "content": user_msg})
    return _complete(messages)


def _edge_tts(text: str, gender: str) -> bytes:
    """Бесплатная озвучка через Microsoft Edge TTS (русские нейроголоса, без ключа)."""
    import asyncio
    import edge_tts
    voice = settings.EDGE_TTS_VOICE_MALE if gender == "male" else settings.EDGE_TTS_VOICE

    async def _gen() -> bytes:
        buf = bytearray()
        communicate = edge_tts.Communicate(text, voice)
        async for chunk in communicate.stream():
            if chunk["type"] == "audio":
                buf += chunk["data"]
        return bytes(buf)

    audio = asyncio.run(_gen())
    if not audio:
        raise RuntimeError("Edge TTS вернул пустой ответ")
    return audio


def _openai_tts(text: str, gender: str) -> bytes:
    import httpx
    voice = settings.OPENAI_TTS_VOICE_MALE if gender == "male" else settings.OPENAI_TTS_VOICE
    resp = httpx.post(
        "https://api.openai.com/v1/audio/speech",
        headers={"Authorization": f"Bearer {settings.OPENAI_API_KEY}"},
        json={"model": settings.OPENAI_TTS_MODEL, "voice": voice, "input": text, "response_format": "mp3"},
        timeout=30,
    )
    if resp.status_code != 200:
        raise RuntimeError(f"OpenAI TTS error {resp.status_code}: {resp.text[:200]}")
    return resp.content


def _resemble_tts(text: str, gender: str) -> bytes:
    """Resemble.ai TTS. Голоса: Varvara (жен) / Михаил (муж). Ответ — base64 mp3."""
    import base64
    import httpx
    voice = settings.RESEMBLE_VOICE_ID_MALE if gender == "male" else settings.RESEMBLE_VOICE_ID
    if not voice:
        raise RuntimeError("Resemble voice_uuid не задан")
    resp = httpx.post(
        "https://f.cluster.resemble.ai/synthesize",
        headers={"Authorization": f"Bearer {settings.RESEMBLE_API_KEY}", "Content-Type": "application/json"},
        json={
            "voice_uuid": voice,
            "data": text,
            "output_format": "mp3",
            "sample_rate": settings.RESEMBLE_SAMPLE_RATE,
        },
        timeout=30,
    )
    if resp.status_code != 200:
        raise RuntimeError(f"Resemble error {resp.status_code}: {resp.text[:200]}")
    body = resp.json()
    audio_b64 = body.get("audio_content")
    if not body.get("success") or not audio_b64:
        raise RuntimeError(f"Resemble: пустой ответ {str(body)[:200]}")
    return base64.b64decode(audio_b64)


def text_to_speech(text: str, gender: str = "female") -> bytes:
    """Озвучка текста голосом Барашка. Приоритет: OpenAI TTS → ElevenLabs.

    Если ни один ключ не задан/сервис недоступен — бросает исключение,
    чтобы фронт откатился на браузерный голос."""
    import re as _re
    # лёгкая чистка (эмодзи/markdown/разделители тысяч) и для OpenAI, и для ElevenLabs
    clean = _re.sub(r"[\U0001F000-\U0001FAFF☀-➿️‍]", "", text)
    clean = _re.sub(r"[*_`#>$₽¥€]", "", clean)
    clean = _re.sub(r"(?<=\d)[\s ,](?=\d)", "", clean)
    clean = _re.sub(r"\s+", " ", clean).strip()[:800]
    if not clean:
        raise ValueError("Пустой текст")

    # 1) Resemble.ai — основной голос Барашка (Varvara/Михаил), если задан ключ
    if settings.RESEMBLE_API_KEY:
        try:
            return _resemble_tts(clean, gender)
        except Exception as e:  # noqa: BLE001 — при сбое падаем на следующий провайдер
            logger.warning("Resemble TTS failed: %s", e)

    # 2) Edge TTS — бесплатный фолбэк (русский Svetlana)
    if settings.EDGE_TTS_ENABLED:
        try:
            return _edge_tts(clean, gender)
        except Exception as e:  # noqa: BLE001 — при сбое падаем на следующий провайдер
            logger.warning("Edge TTS failed: %s", e)

    # 2) OpenAI TTS (если задан ключ)
    if settings.OPENAI_API_KEY:
        return _openai_tts(clean, gender)

    # 3) ElevenLabs
    if not settings.ELEVENLABS_API_KEY:
        raise ValueError("Нет доступного TTS-провайдера")
    voice_id = settings.ELEVENLABS_VOICE_ID_MALE if gender == "male" else settings.ELEVENLABS_VOICE_ID
    text = clean

    import re
    import httpx

    # Чистим: эмодзи, markdown, валюта; убираем разделители тысяч ("25 000" → "25000")
    t = re.sub(r"[\U0001F000-\U0001FAFF☀-➿️‍]", "", text)
    t = re.sub(r"[*_`#>$₽¥€]", "", t)
    t = re.sub(r"(?<=\d)[\s ,](?=\d)", "", t)
    t = re.sub(r"\s+", " ", t).strip()
    if not t:
        raise ValueError("Пустой текст")
    t = t[:800]

    url = f"https://api.elevenlabs.io/v1/text-to-speech/{voice_id}"
    resp = httpx.post(
        url,
        headers={"xi-api-key": settings.ELEVENLABS_API_KEY, "accept": "audio/mpeg"},
        json={
            "text": t,
            "model_id": settings.ELEVENLABS_MODEL,
            "voice_settings": {"stability": 0.4, "similarity_boost": 0.85, "style": 0.35, "use_speaker_boost": True},
        },
        params={"output_format": "mp3_44100_128"},
        timeout=30,
    )
    if resp.status_code != 200:
        raise RuntimeError(f"ElevenLabs error {resp.status_code}: {resp.text[:200]}")
    return resp.content


def interview_coach(db: Session, user: User, message: str, history: list) -> str:
    """Барашек проводит мок-интервью для поступления: вопрос → фидбэк → следующий вопрос."""
    context = _build_student_context(db, user)
    system = (
        "Ты — Барашек 🐑, добрый тренер по собеседованиям для поступления в зарубежные вузы. "
        "Проводишь дружелюбное мок-интервью.\n"
        "ПРАВИЛА:\n"
        "- Задавай по ОДНОМУ вопросу за раз (мотивация, цели, почему этот вуз и специальность, "
        "сильные стороны, планы на будущее, преодоление трудностей).\n"
        "- Если история пустая — тепло поприветствуй и задай ПЕРВЫЙ вопрос.\n"
        "- После ответа студента: 1-2 предложения тёплого фидбэка (что хорошо + что улучшить), "
        "затем следующий вопрос.\n"
        "- Сам вопрос выделяй **жирным**. Коротко, по-русски, поддерживающе, без занудства.\n"
        f"\nПрофиль студента:\n{context}"
    )
    messages = [{"role": "system", "content": system}]
    for h in history[-10:]:
        if h.get("role") in ("user", "assistant"):
            messages.append({"role": h["role"], "content": h["content"]})
    messages.append({"role": "user", "content": message or "Начнём интервью."})
    return _complete(messages)


def barashek_next_action(db: Session, user: User) -> dict:
    """Одно конкретное следующее действие для студента (детерминированно, без Groq).

    Возвращает {text, action_label, action_href, mood, emoji}."""
    from app.models.lesson import Lesson
    from app.models.lesson_view import LessonView

    def res(text, label, href, mood="talking", emoji="🎯"):
        return {"text": text, "action_label": label, "action_href": href, "mood": mood, "emoji": emoji}

    # 1. Профиль (нужен GPA для подбора)
    if not user.gpa:
        return res("Сначала заполни профиль — укажи GPA и баллы тестов. Без этого я не подберу вузы точно.",
                   "Открыть профиль", "/dashboard/profile", "talking", "📝")

    # 2. Выбран ли университет
    progress = db.query(StudentProgress).filter(
        StudentProgress.user_id == user.id,
        StudentProgress.status == ProgressStatus.in_progress,
    ).first()
    if not progress:
        return res("Давай выберем университет — с него начнётся твой путь! Жми «Подобрать по моим данным».",
                   "К университетам", "/dashboard/universities", "talking", "🏛️")

    stage = db.get(Stage, progress.current_stage_id) if progress.current_stage_id else None
    if not stage:
        return res("Ты прошёл все этапы — поздравляю! Горжусь тобой!",
                   "Мой путь", "/dashboard/training", "happy", "🎓")

    # 3. Непросмотренные уроки этапа
    lessons = db.query(Lesson).filter(Lesson.stage_id == stage.id).order_by(Lesson.order).all()
    if lessons:
        viewed = {row[0] for row in db.query(LessonView.lesson_id).filter(
            LessonView.user_id == user.id,
            LessonView.lesson_id.in_([l.id for l in lessons]),
        ).all()}
        unseen = [l for l in lessons if l.id not in viewed]
        if unseen:
            return res(f"Посмотри урок «{unseen[0].title}» — это откроет следующий шаг этапа «{stage.name}».",
                       "Открыть урок", f"/dashboard/lessons/{unseen[0].id}", "talking", "🎬")

    # 4. Невыполненные обязательные требования
    reqs = db.query(Requirement).filter(Requirement.stage_id == stage.id).all()
    done = {row[0] for row in db.query(StudentRequirement.requirement_id).filter(
        StudentRequirement.student_progress_id == progress.id,
        StudentRequirement.completed == True,  # noqa: E712
    ).all()}
    pending = [r for r in reqs if r.id not in done and r.required]

    # дедлайн этапа
    ssd = db.query(StudentStageDeadline).filter(
        StudentStageDeadline.student_progress_id == progress.id,
        StudentStageDeadline.stage_id == stage.id,
    ).first()
    overdue = bool(ssd and (ssd.deadline - date.today()).days < 0)

    if pending:
        mood = "sad" if overdue else "talking"
        urg = "⚠️ Срок уже истёк! " if overdue else ""
        return res(f"{urg}Выполни требование «{pending[0].name}» на этапе «{stage.name}».",
                   "К моему пути", "/dashboard/training", mood, "✅")

    # 5. Всё выполнено — переходи дальше
    return res(f"На этапе «{stage.name}» всё готово — переходи на следующий этап! Ты молодец!",
               "К моему пути", "/dashboard/training", "happy", "🚀")


def barashek_personal_tip(db: Session, user: User) -> str:
    """Короткая тёплая персональная подсказка от Барашка по профилю студента."""
    context = _build_student_context(db, user)
    system = (
        "Ты — Барашек 🐑, тёплый AI-помощник EduBridge. По данным студента дай ОДНУ "
        "короткую (1-2 предложения) персональную, ободряющую подсказку: что ему сделать "
        "следующим шагом, с конкретикой по его профилю. Тепло, по имени, с эмодзи, по-русски. "
        "Только текст подсказки, без префиксов."
    )
    messages = [
        {"role": "system", "content": system},
        {"role": "user", "content": f"Профиль студента:\n{context}"},
    ]
    return _complete(messages, max_tokens=200)


# ── Feature 1: Student chat ───────────────────────────────────────────────────

def student_chat(db: Session, user: User, message: str, history: list) -> str:
    context = _build_student_context(db, user)
    system = (
        "Ты AI-ассистент платформы EduBridge University — помогаешь студентам с поступлением в зарубежные университеты.\n"
        "Отвечай только на русском языке. Будь конкретным, дружелюбным и полезным.\n"
        "Если студент спрашивает о требованиях — ссылайся на его текущий этап.\n\n"
        f"Информация о студенте:\n{context}"
    )
    messages = [{"role": "system", "content": system}]
    for h in history[-10:]:
        if h.get("role") in ("user", "assistant"):
            messages.append({"role": h["role"], "content": h["content"]})
    messages.append({"role": "user", "content": message})
    return _complete(messages)


# ── Onboarding: персонаж «Барашек» ───────────────────────────────────────────

# Поля, которые Барашек собирает в диалоге. Значение — человекочитаемое описание
# для модели + способ, которым его нужно вернуть в profile_updates.
ONBOARDING_FIELDS = {
    "date_of_birth": "дата рождения (формат YYYY-MM-DD)",
    "citizenship": "гражданство / страна (строка)",
    "gpa": "средний балл аттестата (число, например 4.5 или 85)",
    "desired_specialty": "желаемая специальность / направление обучения (строка)",
    "program_level": "уровень программы: одно из bachelor / master / language",
    "country_preference": "предпочитаемые страны обучения (массив строк)",
    "ielts_score": "балл IELTS (число 0-9) либо null если не сдавал",
    "toefl_score": "балл TOEFL (целое 0-120) либо null если не сдавал",
    "sat_score": "балл SAT (целое 0-1600) либо null если не сдавал",
    "hsk_level": "уровень HSK по китайскому (целое 1-6) либо null если не сдавал",
    "max_budget_rmb": "максимальный бюджет на обучение в RMB/год (целое) либо null",
    "wants_language_year": "нужен ли языковой год: одно из yes / no / maybe",
    "preferred_difficulty": "желаемая сложность поступления: одно из Легко / Средний / Сложно",
    "achievements": "достижения, награды, опыт (строка) либо null",
}

# Какие поля обязательны, чтобы считать онбординг завершённым.
ONBOARDING_REQUIRED = [
    "citizenship", "gpa", "desired_specialty", "program_level",
    "country_preference", "max_budget_rmb",
]

# Порядок, в котором Барашек задаёт вопросы.
ONBOARDING_ASK_ORDER = [
    "date_of_birth", "citizenship", "gpa", "desired_specialty",
    "program_level", "country_preference", "max_budget_rmb",
    "ielts_score", "preferred_difficulty",
]

# Тип поля для интерфейса: text | number | date | choice
ONBOARDING_FIELD_TYPES = {
    "date_of_birth": "date", "citizenship": "text", "gpa": "number",
    "desired_specialty": "text", "program_level": "choice",
    "country_preference": "choice", "max_budget_rmb": "number",
    "ielts_score": "number", "toefl_score": "number", "sat_score": "number",
    "hsk_level": "choice", "wants_language_year": "choice",
    "preferred_difficulty": "choice", "achievements": "text",
}

# Поля, которые можно пропустить кнопкой «Не сдавал/Пропустить».
ONBOARDING_SKIPPABLE = {"ielts_score", "toefl_score", "sat_score", "hsk_level", "achievements", "date_of_birth"}

# Шаблонные вопросы — подставляются, если модель забыла задать вопрос.
ONBOARDING_QUESTIONS = {
    "date_of_birth": "Какая у тебя дата рождения? (ГГГГ-ММ-ДД)",
    "citizenship": "Из какой ты страны?",
    "gpa": "Какой у тебя средний балл аттестата (GPA)?",
    "desired_specialty": "Кем хочешь стать — какая специальность тебе интересна?",
    "program_level": "Какой уровень обучения тебе нужен?",
    "country_preference": "В какой стране хочешь учиться?",
    "max_budget_rmb": "Какой у тебя бюджет на обучение в год (в RMB)?",
    "ielts_score": "Какой у тебя балл IELTS? Если не сдавал — нажми «Пропустить».",
    "toefl_score": "Какой у тебя балл TOEFL? Если не сдавал — пропусти.",
    "sat_score": "Какой у тебя балл SAT? Если не сдавал — пропусти.",
    "hsk_level": "Какой у тебя уровень HSK?",
    "preferred_difficulty": "Какую сложность поступления предпочитаешь?",
}


def onboarding_field_meta(db: Session, field: str) -> dict:
    """Тип поля и варианты-кнопки для текущего вопроса онбординга."""
    if not field or field not in ONBOARDING_FIELDS:
        return {"field": "", "type": "text", "options": [], "skippable": False}
    options: list = []
    if field == "program_level":
        options = [
            {"label": "Бакалавриат", "value": "bachelor"},
            {"label": "Магистратура", "value": "master"},
            {"label": "Языковой год", "value": "language"},
        ]
    elif field == "wants_language_year":
        options = [{"label": "Да", "value": "yes"}, {"label": "Нет", "value": "no"}, {"label": "Возможно", "value": "maybe"}]
    elif field == "preferred_difficulty":
        options = [{"label": "Легко", "value": "Легко"}, {"label": "Средний", "value": "Средний"}, {"label": "Сложно", "value": "Сложно"}]
    elif field == "hsk_level":
        options = [{"label": str(i), "value": str(i)} for i in range(1, 7)]
    elif field == "country_preference":
        options = [{"label": c, "value": c} for c in _system_countries(db)]
    return {
        "field": field,
        "type": ONBOARDING_FIELD_TYPES.get(field, "text"),
        "options": options,
        "skippable": field in ONBOARDING_SKIPPABLE,
    }


def _filled_fields(user: User) -> dict:
    """Текущие заполненные поля профиля, релевантные онбордингу."""
    result = {}
    for field in ONBOARDING_FIELDS:
        val = getattr(user, field, None)
        if val not in (None, "", []):
            result[field] = val
    return result


# Частые русские названия стран → каноничные (как обычно в базе университетов).
_RU_COUNTRY = {
    "китай": "China", "кнр": "China",
    "сша": "USA", "америка": "USA", "соединённые штаты": "USA",
    "германия": "Germany", "италия": "Italy", "франция": "France",
    "испания": "Spain", "великобритания": "UK", "англия": "UK",
    "канада": "Canada", "южная корея": "South Korea", "корея": "South Korea",
    "япония": "Japan", "турция": "Turkey", "россия": "Russia",
    "казахстан": "Kazakhstan", "кыргызстан": "Kyrgyzstan", "киргизия": "Kyrgyzstan",
}


def _system_countries(db: Session) -> list[str]:
    """Список стран, реально присутствующих в каталоге университетов."""
    rows = db.query(University.country).filter(
        University.deleted_at.is_(None), University.country.isnot(None)
    ).distinct().all()
    return [r[0] for r in rows if r[0]]


def _normalize_countries(values: list, system_countries: list[str]) -> list:
    """Приводим страны к написанию, которое есть в системе (чтобы совпадало
    с выпадающим списком профиля и работал подбор). Рус→англ как запасной вариант."""
    lookup = {c.lower(): c for c in system_countries}
    result = []
    for v in values:
        low = str(v).strip().lower()
        if low in lookup:
            result.append(lookup[low])
        elif low in _RU_COUNTRY:
            mapped = _RU_COUNTRY[low]
            result.append(lookup.get(mapped.lower(), mapped))
        else:
            result.append(str(v).strip())
    # уникальные, без пустых, с сохранением порядка
    seen = set()
    return [x for x in result if x and not (x in seen or seen.add(x))]


def onboarding_chat(db: Session, user: User, message: str, history: list, skipped: list | None = None) -> dict:
    """Диалог-онбординг с персонажем «Барашек».

    Барашек тепло встречает студента, благодарит и мягко, но настойчиво
    вытягивает данные профиля прямо из разговора. Возвращает:
      {"reply": str, "profile_updates": dict, "completed": bool}
    """
    is_start = not any(h.get("role") == "assistant" for h in history)
    filled = _filled_fields(user)
    filled_desc = ", ".join(f"{k}={v}" for k, v in filled.items()) or "пока ничего"
    system_countries = _system_countries(db)
    countries_hint = (
        f"country_preference бери ТОЧНО из этих стран системы: {', '.join(system_countries)}.\n"
        if system_countries else ""
    )

    # AI ТОЛЬКО извлекает данные из ответа студента и даёт одно слово похвалы.
    # Сам вопрос и кнопки строит backend детерминированно — чтобы они всегда совпадали.
    system = (
        "Ты — Барашек, тёплый AI-талисман EduBridge. В этом ответе ты ТОЛЬКО извлекаешь "
        "данные из последнего сообщения студента в profile_updates и даёшь ОДНО слово похвалы.\n"
        "Форматы значений:\n"
        "- gpa, ielts_score: число; toefl_score/sat_score/hsk_level/max_budget_rmb: целое\n"
        "- date_of_birth: YYYY-MM-DD; citizenship/desired_specialty/achievements: строка\n"
        "- program_level: bachelor|master|language; wants_language_year: yes|no|maybe\n"
        "- preferred_difficulty: Легко|Средний|Сложно; country_preference: массив строк\n"
        f"{countries_hint}"
        "Если студент ничего конкретного не сообщил, ответил не по теме или «не сдавал/пропустить» — "
        "profile_updates пустой объект {}.\n"
        "praise — одно тёплое слово на русском («Супер!», «Отлично!», «Класс!», «Здорово!»).\n"
        f"Уже известно о студенте: {filled_desc}\n\n"
        'ФОРМАТ — строго JSON: {"profile_updates": {...}, "praise": "Отлично!"}'
    )

    messages = [{"role": "system", "content": system}]
    for h in history[-8:]:
        if h.get("role") in ("user", "assistant"):
            messages.append({"role": h["role"], "content": h["content"]})
    messages.append({"role": "user", "content": message})

    parsed = _parse_json(_complete(messages), {"profile_updates": {}, "praise": "Отлично!"})

    updates = parsed.get("profile_updates")
    if not isinstance(updates, dict):
        updates = {}
    clean_updates = _sanitize_onboarding_updates(updates)
    if "country_preference" in clean_updates:
        clean_updates["country_preference"] = _normalize_countries(
            clean_updates["country_preference"], system_countries
        )

    # Детерминированно вычисляем следующий вопрос (с учётом только что извлечённых полей
    # и пропущенных пользователем). Поле и текст вопроса всегда совпадают → кнопки в
    # интерфейсе соответствуют вопросу.
    filled_after = set(filled) | set(clean_updates) | set(skipped or [])
    next_field = next((f for f in ONBOARDING_ASK_ORDER if f not in filled_after), "")
    completed = next_field == ""

    praise = parsed.get("praise")
    praise = praise.strip() if isinstance(praise, str) and praise.strip() else "Отлично!"

    if completed:
        reply = "Готово! Я узнал всё, что нужно. Теперь подберу тебе университеты — жми кнопку ниже."
    else:
        question = ONBOARDING_QUESTIONS.get(next_field, "Расскажи мне ещё немного о себе.")
        reply = f"Здравствуй, {user.full_name}! **{question}**" if is_start else f"{praise} **{question}**"

    return {
        "reply": reply,
        "profile_updates": clean_updates,
        "next_field": next_field,
        "completed": completed,
    }


def _sanitize_onboarding_updates(updates: dict) -> dict:
    """Оставляем только известные поля и приводим значения к нужным типам."""
    clean: dict = {}
    for key, val in updates.items():
        if key not in ONBOARDING_FIELDS or val in (None, "", []):
            continue
        try:
            if key == "gpa":
                clean[key] = float(val)
            elif key == "ielts_score":
                clean[key] = float(val)
            elif key in ("toefl_score", "sat_score", "hsk_level", "max_budget_rmb"):
                clean[key] = int(float(val))
            elif key == "country_preference":
                if isinstance(val, str):
                    clean[key] = [val]
                elif isinstance(val, list):
                    clean[key] = [str(v) for v in val if v]
            elif key == "program_level":
                if str(val).lower() in ("bachelor", "master", "language"):
                    clean[key] = str(val).lower()
            elif key == "wants_language_year":
                if str(val).lower() in ("yes", "no", "maybe"):
                    clean[key] = str(val).lower()
            elif key == "preferred_difficulty":
                if str(val) in ("Легко", "Средний", "Сложно"):
                    clean[key] = str(val)
            else:  # citizenship, desired_specialty, achievements, date_of_birth
                clean[key] = str(val)
        except (ValueError, TypeError):
            continue
    return clean


# ── Feature 2: Check motivation letter ───────────────────────────────────────

def check_motivation_letter(text: str) -> dict:
    prompt = (
        "Ты эксперт по поступлению в зарубежные университеты. Проверь мотивационное письмо.\n"
        "Ответь строго в формате JSON (только JSON, без markdown):\n"
        '{"score": <1-10>, "strengths": ["..."], "weaknesses": ["..."], "suggestions": ["..."], "verdict": "краткое резюме"}\n\n'
        f"Мотивационное письмо:\n{text}"
    )
    raw = _complete([{"role": "user", "content": prompt}])
    return _parse_json(raw, {
        "score": 0, "strengths": [], "weaknesses": [],
        "suggestions": ["Не удалось проанализировать"], "verdict": raw[:200]
    })


# ── Feature 3: Check document image ──────────────────────────────────────────

def check_document_image(image_bytes: bytes, mime_type: str) -> dict:
    b64 = base64.standard_b64encode(image_bytes).decode()
    messages = [{
        "role": "user",
        "content": [
            {"type": "image_url", "image_url": {"url": f"data:{mime_type};base64,{b64}"}},
            {
                "type": "text",
                "text": (
                    "Проверь этот документ. Ответь строго в формате JSON (без markdown):\n"
                    '{"ok": true/false, "document_type": "тип", "issues": ["..."], '
                    '"empty_fields": ["..."], "verdict": "заключение", "recommendations": ["..."]}'
                ),
            },
        ],
    }]
    raw = _complete(messages, model=VISION_MODEL)
    return _parse_json(raw, {
        "ok": False, "document_type": "Неизвестно", "issues": [raw[:200]],
        "empty_fields": [], "verdict": "Ошибка анализа", "recommendations": []
    })


def _pdf_to_png(content: bytes) -> bytes:
    """Рендер первой страницы PDF в PNG (для vision-проверки аттестатов/писем)."""
    import fitz  # PyMuPDF
    doc = fitz.open(stream=content, filetype="pdf")
    try:
        page = doc.load_page(0)
        pix = page.get_pixmap(dpi=150)
        return pix.tobytes("png")
    finally:
        doc.close()


def _docx_to_text(content: bytes) -> str:
    """Извлечь текст из .docx."""
    import io
    import docx
    d = docx.Document(io.BytesIO(content))
    return "\n".join(p.text for p in d.paragraphs if p.text)


def check_document_text(text: str) -> dict:
    """Проверка текстового документа (письмо/эссе/справка) — честно, но мягко."""
    text = (text or "").strip()
    if not text:
        return {"ok": False, "document_type": "Документ", "issues": ["Не удалось прочитать текст"],
                "empty_fields": [], "verdict": "Файл пустой или нечитаемый", "recommendations": ["Загрузи файл с текстом"]}
    prompt = (
        "Ты эксперт по поступлению в зарубежные вузы. Проверь документ студента "
        "(мотивационное/рекомендательное письмо, эссе, справку) честно, но доброжелательно.\n"
        "Ответь строго JSON (без markdown):\n"
        '{"ok": true/false, "document_type": "тип", "issues": ["..."], '
        '"empty_fields": ["..."], "verdict": "краткое заключение", "recommendations": ["..."]}\n\n'
        f"Текст документа:\n{text[:6000]}"
    )
    raw = _complete([{"role": "user", "content": prompt}])
    return _parse_json(raw, {
        "ok": False, "document_type": "Документ", "issues": [raw[:200]],
        "empty_fields": [], "verdict": "Не удалось проанализировать", "recommendations": []
    })


def check_document_file(content: bytes, mime: str = "", filename: str = "") -> dict:
    """Проверка документа Барашком: фото (JPG/PNG/WEBP), PDF, DOCX, TXT."""
    low_mime = (mime or "").lower()
    low_name = (filename or "").lower()
    if "pdf" in low_mime or low_name.endswith(".pdf"):
        return check_document_image(_pdf_to_png(content), "image/png")
    if "wordprocessingml" in low_mime or low_name.endswith(".docx"):
        return check_document_text(_docx_to_text(content))
    if low_mime.startswith("text/") or low_name.endswith(".txt"):
        return check_document_text(content.decode("utf-8", errors="ignore"))
    return check_document_image(content, mime or "image/jpeg")


# ── Feature 3b: Проверка документа с учётом ЗАЯВЛЕННОГО типа ──────────────────
# Студент указывает, что именно он загружает (паспорт / аттестат / диплом …),
# а ИИ сверяет реальное содержимое файла с этим типом и решает, принять ли его.

DOCUMENT_TYPE_LABELS: dict[str, str] = {
    "passport": "Паспорт / удостоверение личности",
    "certificate": "Аттестат о среднем образовании",
    "diploma": "Диплом о высшем образовании",
    "transcript": "Транскрипт (выписка с оценками)",
    "language_certificate": "Языковой сертификат (IELTS / TOEFL / HSK)",
    "motivation_letter": "Мотивационное письмо",
    "recommendation": "Рекомендательное письмо",
    "photo": "Фото на документы",
    "medical": "Медицинская справка",
    "other": "Прочий документ",
}


def _doc_meta_line(mime: str, filename: str, size: int | None) -> str:
    parts = []
    if filename:
        parts.append(f"имя файла: {filename}")
    if mime:
        parts.append(f"формат (MIME): {mime}")
    if size:
        parts.append(f"размер: {size // 1024} КБ" if size >= 1024 else f"размер: {size} Б")
    return "; ".join(parts) or "метаданные недоступны"


def _verify_instruction(expected_label: str, meta: str) -> str:
    return (
        "Ты строгий проверяющий документов для поступления в зарубежные вузы. "
        f"Студент УТВЕРЖДАЕТ, что это документ типа: «{expected_label}».\n"
        f"Метаданные файла: {meta}.\n\n"
        "Проверь РЕАЛЬНОЕ содержимое файла:\n"
        "1) Определи фактический тип документа по тому, что видишь/читаешь.\n"
        "2) Совпадает ли он с заявленным типом? Если загружен, например, аттестат "
        "вместо паспорта — это НЕСООТВЕТСТВИЕ, принимать НЕЛЬЗЯ.\n"
        "3) Документ читаемый (не размытый, не пустой, не обрезанный)?\n"
        "4) Заполнены ли ключевые поля, характерные для этого типа документа?\n"
        "НЕ одобряй документ только из-за факта загрузки. Одобряй ТОЛЬКО когда тип "
        "совпадает И документ читаем И ключевые поля на месте.\n"
        "Пиши КРАТКО, как учитель: без длинных объяснений и повторов.\n\n"
        "Ответь строго в формате JSON (без markdown):\n"
        '{"detected_type": "фактический тип, 2-4 слова", '
        '"type_match": true/false, "readable": true/false, '
        '"reasons": ["1-2 коротких пункта"], '
        '"issues": ["короткие проблемы, если есть"], '
        '"recommendations": ["коротко, что исправить"], '
        '"verdict": "одно короткое предложение"}'
    )


def _vision_complete(image_bytes: bytes, mime_type: str, instruction: str) -> str:
    b64 = base64.standard_b64encode(image_bytes).decode()
    messages = [{
        "role": "user",
        "content": [
            {"type": "image_url", "image_url": {"url": f"data:{mime_type};base64,{b64}"}},
            {"type": "text", "text": instruction},
        ],
    }]
    return _complete(messages, model=VISION_MODEL)


def _finalize_verdict(parsed: dict, expected_label: str) -> dict:
    type_match = bool(parsed.get("type_match"))
    readable = parsed.get("readable")
    readable = True if readable is None else bool(readable)
    issues = parsed.get("issues") or []
    reasons = parsed.get("reasons") or []
    detected = parsed.get("detected_type") or "Неизвестно"

    # Одобряем ТОЛЬКО при совпадении типа, читаемости и отсутствии проблем.
    ok = type_match and readable and len(issues) == 0

    # Жёсткое правило: несоответствие типа = отклонение с явной причиной.
    if not type_match:
        msg = f"Тип не совпадает: ожидался «{expected_label}», распознано «{detected}»."
        if msg not in reasons:
            reasons = [msg] + reasons
        ok = False
    if not readable and "Документ нечитаемый или пустой" not in reasons:
        reasons = reasons + ["Документ нечитаемый или пустой"]

    return {
        "ok": ok,
        "status": "approved" if ok else "rejected",
        "expected_type_label": expected_label,
        "detected_type": detected,
        "type_match": type_match,
        "readable": readable,
        "reasons": reasons,
        "issues": issues,
        "recommendations": parsed.get("recommendations") or [],
        "verdict": parsed.get("verdict") or ("Документ принят" if ok else "Документ отклонён"),
    }


def verify_document(
    content: bytes,
    mime: str = "",
    filename: str = "",
    expected_type: str = "other",
    file_size: int | None = None,
    expected_label: str | None = None,
) -> dict:
    """Проверка документа с учётом ЗАЯВЛЕННОГО типа.

    Решение принимается на основе реальных данных: содержимого файла, его размера,
    формата и метаданных. Возвращает вердикт (ok), статус, распознанный тип,
    флаг совпадения типа и СПИСОК ПРИЧИН принятия/отклонения.

    `expected_label` — произвольное название ожидаемого документа (напр. название
    требования этапа «Академическая справка с оценками»); если не задано — берётся
    человекочитаемая метка по `expected_type`."""
    label = expected_label or DOCUMENT_TYPE_LABELS.get(expected_type, DOCUMENT_TYPE_LABELS["other"])
    size = file_size if file_size is not None else len(content)
    meta = _doc_meta_line(mime, filename, size)
    instruction = _verify_instruction(label, meta)

    low_mime = (mime or "").lower()
    low_name = (filename or "").lower()

    fallback = {
        "detected_type": "Неизвестно", "type_match": False, "readable": False,
        "reasons": ["Не удалось проанализировать документ"], "issues": [],
        "recommendations": ["Загрузите более чёткий файл (JPG/PNG/PDF)"],
        "verdict": "Не удалось проверить",
    }

    try:
        if "pdf" in low_mime or low_name.endswith(".pdf"):
            raw = _vision_complete(_pdf_to_png(content), "image/png", instruction)
        elif "wordprocessingml" in low_mime or low_name.endswith(".docx"):
            text = _docx_to_text(content)
            raw = _complete([{"role": "user", "content": instruction + f"\n\nТекст документа:\n{text[:6000]}"}])
        elif low_mime.startswith("text/") or low_name.endswith(".txt"):
            text = content.decode("utf-8", errors="ignore")
            raw = _complete([{"role": "user", "content": instruction + f"\n\nТекст документа:\n{text[:6000]}"}])
        else:
            raw = _vision_complete(content, mime or "image/jpeg", instruction)
    except Exception:
        logger.exception("verify_document failed")
        return _finalize_verdict(fallback, label)

    parsed = _parse_json(raw, fallback)
    return _finalize_verdict(parsed, label)


# ── Feature 4: Extract profile from document image ───────────────────────────

def extract_profile_from_document(image_bytes: bytes, mime_type: str) -> dict:
    b64 = base64.standard_b64encode(image_bytes).decode()
    messages = [{
        "role": "user",
        "content": [
            {"type": "image_url", "image_url": {"url": f"data:{mime_type};base64,{b64}"}},
            {
                "type": "text",
                "text": (
                    "Извлеки данные из этого документа. Ответь строго в формате JSON (без markdown):\n"
                    '{"full_name": "...", "date_of_birth": "YYYY-MM-DD или null", '
                    '"citizenship": "...", "gpa": null, "ielts_score": null, "sat_score": null}'
                    "\nЕсли поле не найдено — null."
                ),
            },
        ],
    }]
    raw = _complete(messages, model=VISION_MODEL)
    return _parse_json(raw, {})


# ── Feature 5: University recommendations ────────────────────────────────────

def recommend_universities(db: Session, user: User) -> list:
    unis = db.query(University).filter(University.deleted_at.is_(None)).limit(30).all()

    def _uni_line(u: University) -> str:
        specs_cn = ", ".join(u.programs_bachelor_chinese or []) if u.programs_bachelor_chinese else ""
        specs_en = ", ".join(u.programs_bachelor_english or []) if u.programs_bachelor_english else ""
        specs_old = ", ".join(u.specialties) if u.specialties else ""
        specs = specs_cn or specs_en or specs_old or "не указаны"

        tuition = u.tuition_bachelor or u.cost
        budget_ok = ""
        if user.max_budget_rmb and tuition:
            budget_ok = " [БЮДЖЕТ ОК]" if tuition <= user.max_budget_rmb else " [ДОРОГО]"

        return (
            f"- {u.name} ({u.country}, {u.city})"
            f", сложность: {u.difficulty or '?'}"
            f", рейтинг: {u.rating or '?'}"
            f", стоимость бакалавриат: {tuition or '?'} RMB/год{budget_ok}"
            f", языковой год: {'да' if u.has_language_year else 'нет'}"
            f", дедлайн: {u.deadline or '?'}"
            f", специальности: {specs[:120]}"
            f", мин. требования: {(u.min_requirements or '')[:80]}"
        )

    uni_list = "\n".join(_uni_line(u) for u in unis)
    context = _build_student_context(db, user)
    prompt = (
        "Подбери топ-5 подходящих университетов для студента с учётом его бюджета, уровня программы, "
        "желаемой специальности и предпочитаемой сложности. "
        "Ответь строго в формате JSON-массив (без markdown):\n"
        '[{"name": "...", "match_percent": 85, "reasons": ["..."], "concerns": ["..."]}]\n\n'
        f"Профиль студента:\n{context}\n\nДоступные университеты:\n{uni_list}"
    )
    raw = _complete([{"role": "user", "content": prompt}])
    return _parse_json_list(raw)


# ── Feature 6: Translate motivation letter ────────────────────────────────────

def translate_letter(text: str, target_language: str = "английский") -> str:
    prompt = (
        f"Переведи это мотивационное письмо на {target_language}. "
        "Сохрани стиль и структуру. Верни только перевод без пояснений.\n\n"
        f"{text}"
    )
    return _complete([{"role": "user", "content": prompt}], max_tokens=3000)


# ── Feature 7: Smart reminders ────────────────────────────────────────────────

def get_smart_reminders(db: Session, user: User) -> list:
    context = _build_student_context(db, user)
    prompt = (
        "На основе статуса студента сформулируй 2-4 персональных напоминания. "
        "Ответь строго в формате JSON-массив (без markdown):\n"
        '[{"priority": "high/medium/low", "icon": "📋", "message": "текст"}]\n\n'
        f"Статус студента:\n{context}"
    )
    raw = _complete([{"role": "user", "content": prompt}])
    result = _parse_json_list(raw)
    if not result:
        return [{"priority": "medium", "icon": "ℹ️", "message": "Проверьте ваш прогресс в личном кабинете"}]
    return result


# ── Feature 8: Admin — student summary ───────────────────────────────────────

def get_student_summary(db: Session, user_id: str) -> dict:
    user = db.get(User, uuid.UUID(user_id))
    if not user:
        return {"summary": "Студент не найден", "action_items": [], "risk_level": "low", "recommendation": ""}
    context = _build_student_context(db, user)
    prompt = (
        "Ты менеджер образовательного агентства. Составь краткую сводку по студенту для коллеги.\n"
        "Ответь строго в формате JSON (без markdown):\n"
        '{"summary": "3-4 предложения", "action_items": ["..."], '
        '"risk_level": "low/medium/high", "recommendation": "главная рекомендация"}\n\n'
        f"Данные:\n{context}"
    )
    raw = _complete([{"role": "user", "content": prompt}])
    return _parse_json(raw, {
        "summary": raw[:300], "action_items": [], "risk_level": "medium", "recommendation": ""
    })


# ── Feature 10: Admin — analytics ────────────────────────────────────────────

def answer_analytics_query(db: Session, question: str) -> dict:
    from sqlalchemy import func
    from app.models.lead import Lead

    student_filter = (User.role == "student", User.deleted_at.is_(None))
    total_students = db.query(func.count(User.id)).filter(*student_filter).scalar() or 0
    active = db.query(func.count(User.id)).filter(
        *student_filter, User.account_status == AccountStatus.active
    ).scalar() or 0
    enrolled = db.query(func.count(User.id)).filter(
        *student_filter, User.account_status == AccountStatus.enrolled
    ).scalar() or 0
    archived = db.query(func.count(User.id)).filter(
        *student_filter, User.account_status == AccountStatus.archived
    ).scalar() or 0
    total_leads = db.query(func.count(Lead.id)).scalar() or 0

    avg_gpa = round(db.query(func.avg(User.gpa)).filter(*student_filter).scalar() or 0, 2)
    avg_ielts = round(db.query(func.avg(User.ielts_score)).filter(*student_filter).scalar() or 0, 1)

    progresses = db.query(StudentProgress).filter(StudentProgress.status == ProgressStatus.in_progress).all()

    # Batch-load related stages/universities to avoid N+1
    stage_ids = {p.current_stage_id for p in progresses if p.current_stage_id}
    uni_ids = {p.university_id for p in progresses if p.university_id}
    stage_names = {s.id: s.name for s in db.query(Stage).filter(Stage.id.in_(stage_ids)).all()} if stage_ids else {}
    uni_names = {u.id: u.name for u in db.query(University).filter(University.id.in_(uni_ids)).all()} if uni_ids else {}

    today = date.today()
    overdue_pairs = {
        (d.student_progress_id, d.stage_id)
        for d in db.query(StudentStageDeadline).filter(StudentStageDeadline.deadline < today).all()
    }

    stage_dist: dict = {}
    uni_students: dict = {}
    overdue = 0
    for p in progresses:
        if p.current_stage_id:
            name = stage_names.get(p.current_stage_id)
            if name:
                stage_dist[name] = stage_dist.get(name, 0) + 1
            if (p.id, p.current_stage_id) in overdue_pairs:
                overdue += 1
        uname = uni_names.get(p.university_id)
        if uname:
            uni_students[uname] = uni_students.get(uname, 0) + 1

    uni_filter = (University.deleted_at.is_(None),)
    unis_count = db.query(func.count(University.id)).filter(*uni_filter).scalar() or 0

    uni_by_country = {
        (country or "не указана"): cnt
        for country, cnt in db.query(University.country, func.count(University.id))
        .filter(*uni_filter)
        .group_by(University.country)
        .order_by(func.count(University.id).desc())
        .all()
    }
    uni_by_difficulty = {
        (diff or "не указана"): cnt
        for diff, cnt in db.query(University.difficulty, func.count(University.id))
        .filter(*uni_filter)
        .group_by(University.difficulty)
        .all()
    }
    uni_with_language_year = db.query(func.count(University.id)).filter(
        *uni_filter, University.has_language_year.is_(True)
    ).scalar() or 0

    db_context = (
        f"Данные системы EduBridge:\n"
        f"- Студентов всего: {total_students} (активных: {active}, поступили: {enrolled}, архив: {archived})\n"
        f"- Заявок (лидов): {total_leads}\n"
        f"- Средний GPA: {avg_gpa}, средний IELTS: {avg_ielts}\n"
        f"- Просроченных дедлайнов: {overdue}\n"
        f"- Распределение по этапам: {stage_dist}\n"
        f"- Студентов по университетам: {uni_students}\n"
        f"- Университетов в базе: {unis_count}\n"
        f"- Университетов по странам: {uni_by_country}\n"
        f"- Университетов по сложности поступления: {uni_by_difficulty}\n"
        f"- Университетов с языковым годом: {uni_with_language_year}"
    )

    is_export = any(w in question.lower() for w in ["экспорт", "excel", "xlsx", "скачать", "выгрузи", "список студентов"])

    prompt = (
        "Ты аналитик образовательной платформы. Ответь на вопрос администратора, используя данные системы.\n"
        "Отвечай по-русски. Будь конкретным, давай цифры и факты.\n\n"
        f"{db_context}\n\nВопрос: {question}"
    )
    answer = _complete([{"role": "user", "content": prompt}])
    return {"answer": answer, "can_export": is_export}


# ── Feature 11: Admin — reminder draft ───────────────────────────────────────

def draft_reminder(db: Session, user_id: str) -> dict:
    user = db.get(User, uuid.UUID(user_id))
    if not user:
        return {"draft": "", "user_name": ""}
    context = _build_student_context(db, user)
    prompt = (
        "Напиши персональное напоминание от менеджера образовательного агентства студенту. "
        "Тон: профессиональный, но дружелюбный. Длина: 2-3 предложения. Только текст сообщения.\n\n"
        f"Данные студента:\n{context}"
    )
    draft = _complete([{"role": "user", "content": prompt}])
    return {"draft": draft, "user_name": user.full_name}


# ── Feature 12: Admission score ───────────────────────────────────────────────

def get_admission_score(db: Session, user_id: str) -> dict:
    user = db.get(User, uuid.UUID(user_id))
    if not user:
        return {"score": 0, "breakdown": {}, "verdict": "Студент не найден", "weak_points": [], "strong_points": [], "recommendations": []}

    gpa_score = _gpa_percent(user.gpa)
    # Best available English score: IELTS or TOEFL
    ielts_score = min(100, int((user.ielts_score or 0) / 9.0 * 100))
    toefl_score_norm = min(100, int((user.toefl_score or 0) / 120 * 100)) if user.toefl_score else 0
    english_score = max(ielts_score, toefl_score_norm)
    sat_score = min(100, int((user.sat_score or 0) / 1600 * 100)) if user.sat_score else 0
    hsk_score = min(100, int((user.hsk_level or 0) / 6 * 100)) if user.hsk_level else 0
    # Boost from HSK (Chinese universities bonus)
    lang_bonus = max(english_score, hsk_score)
    profile_fields = [user.gpa, user.ielts_score or user.toefl_score, user.citizenship,
                      user.desired_specialty, user.country_preference, user.achievements,
                      user.email, user.phone, user.program_level, user.max_budget_rmb]
    completeness = int(sum(1 for f in profile_fields if f) / len(profile_fields) * 100)
    overall = int(gpa_score * 0.30 + lang_bonus * 0.30 + sat_score * 0.15 + hsk_score * 0.10 + completeness * 0.15)

    context = _build_student_context(db, user)
    prompt = (
        "Оцени шансы студента на поступление в зарубежный университет. "
        "Ответь строго в формате JSON (без markdown):\n"
        '{"verdict": "2-3 предложения", "weak_points": ["..."], "strong_points": ["..."], "recommendations": ["..."]}\n\n'
        f"Данные:\n{context}\n"
        f"Расчётный балл: {overall}/100 (GPA: {gpa_score}/100, Английский: {english_score}/100, "
        f"HSK: {hsk_score}/100, SAT: {sat_score}/100, профиль: {completeness}/100)"
    )
    raw = _complete([{"role": "user", "content": prompt}])
    ai = _parse_json(raw, {"verdict": raw[:300], "weak_points": [], "strong_points": [], "recommendations": []})

    return {
        "score": overall,
        "breakdown": {
            "gpa": gpa_score,
            "english": english_score,
            "hsk": hsk_score,
            "sat": sat_score,
            "completeness": completeness,
        },
        "verdict": ai.get("verdict", ""),
        "weak_points": ai.get("weak_points", []),
        "strong_points": ai.get("strong_points", []),
        "recommendations": ai.get("recommendations", []),
    }


# ── Compare universities with AI ─────────────────────────────────────────────

def compare_universities(db: Session, user: User, university_ids: list[str]) -> dict:
    unis = [db.get(University, uuid.UUID(uid)) for uid in university_ids]
    unis = [u for u in unis if u]

    uni_blocks = []
    for u in unis:
        specs_cn = ", ".join(u.programs_bachelor_chinese or []) if u.programs_bachelor_chinese else ""
        specs_en = ", ".join(u.programs_bachelor_english or []) if u.programs_bachelor_english else ""
        specs_old = ", ".join(u.specialties) if u.specialties else ""
        specs = specs_cn or specs_en or specs_old or "не указаны"
        docs = ", ".join(u.documents_bachelor or []) if u.documents_bachelor else "не указаны"
        uni_blocks.append(
            f"- {u.name} ({u.country}, {u.city})\n"
            f"  Рейтинг: {u.rating or 'нет'}, Сложность: {u.difficulty or 'не указана'}\n"
            f"  Стоимость бакалавриат: {u.tuition_bachelor or u.cost or 'нет'} RMB/год"
            f"{f', магистратура: {u.tuition_masters} RMB/год' if u.tuition_masters else ''}\n"
            f"  Языковой год: {'доступен' if u.has_language_year else 'нет'}"
            f"{f', стоимость: {u.tuition_language_year} RMB/год' if u.tuition_language_year else ''}\n"
            f"  Дедлайн: {u.deadline or 'не указан'}\n"
            f"  Мин. требования: {(u.min_requirements or 'не указаны')[:150]}\n"
            f"  Специальности: {specs[:150]}\n"
            f"  Документы: {docs[:150]}\n"
            f"  Общежитие: {(u.dormitory_info or 'не указано')[:100]}\n"
            f"  Описание: {(u.description or '')[:150]}"
        )

    student_ctx = _build_student_context(db, user)

    prompt = (
        "Ты эксперт по поступлению в зарубежные университеты. "
        "Сравни университеты и дай персональный анализ для конкретного студента.\n"
        "Ответь строго в формате JSON (без markdown):\n"
        "{\n"
        '  "winner": "название лучшего университета для этого студента",\n'
        '  "winner_reason": "2-3 предложения почему именно он",\n'
        '  "per_university": [\n'
        '    {"name": "...", "pros": ["..."], "cons": ["..."], "admission_chance": "высокие/средние/низкие", "fit_score": 85}\n'
        "  ],\n"
        '  "advice": "главный совет студенту по выбору",\n'
        '  "red_flags": ["риск 1", "риск 2"]\n'
        "}\n\n"
        f"Профиль студента:\n{student_ctx}\n\n"
        f"Сравниваемые университеты:\n" + "\n".join(uni_blocks)
    )

    raw = _complete([{"role": "user", "content": prompt}], max_tokens=2048)
    return _parse_json(raw, {
        "winner": "",
        "winner_reason": raw[:300],
        "per_university": [],
        "advice": "",
        "red_flags": [],
    })


# ── Admin Excel export ────────────────────────────────────────────────────────

def export_students_excel(db: Session) -> bytes:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Студенты"

    headers = ["ФИО", "Логин", "Email", "Телефон", "GPA", "IELTS", "SAT",
               "Статус", "Специальность", "Гражданство", "Дата регистрации"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = openpyxl.styles.Font(bold=True)

    students = db.query(User).filter(User.role == "student", User.deleted_at.is_(None)).all()
    for row, s in enumerate(students, 2):
        progress = db.query(StudentProgress).filter(
            StudentProgress.user_id == s.id,
            StudentProgress.status == ProgressStatus.in_progress,
        ).first()
        stage_name = ""
        if progress and progress.current_stage_id:
            stage = db.get(Stage, progress.current_stage_id)
            stage_name = stage.name if stage else ""

        ws.cell(row=row, column=1, value=s.full_name)
        ws.cell(row=row, column=2, value=s.login)
        ws.cell(row=row, column=3, value=s.email)
        ws.cell(row=row, column=4, value=s.phone or "")
        ws.cell(row=row, column=5, value=s.gpa or "")
        ws.cell(row=row, column=6, value=s.ielts_score or "")
        ws.cell(row=row, column=7, value=s.sat_score or "")
        ws.cell(row=row, column=8, value=s.account_status.value)
        ws.cell(row=row, column=9, value=s.desired_specialty or "")
        ws.cell(row=row, column=10, value=s.citizenship or "")
        ws.cell(row=row, column=11, value=s.created_at.strftime("%d.%m.%Y") if s.created_at else "")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
